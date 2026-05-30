#!/usr/bin/env python3
"""
ground_covered_leaderboard.py  —  League-wide "secondary distance" steal leaderboard
====================================================================================

The thesis of this project: slow runners like **Josh Naylor** (sprint 24.4 ft/s, ~2nd
speed percentile, yet 30 SB / 2 CS in 2025) and **Juan Soto** steal bases not on raw
speed but on the **ground they cover between the pitcher's first move and pitch release**
— a big jump / good timing, not a fast 90 ft.

That "ground covered" is **already a native Statcast metric** — no computer vision needed:

    gain_to_release_ft = r_secondary_lead - r_primary_lead      (Baseball Savant)
                       = (lead at pitch release) - (lead at pitcher's first move)

`fetch_leads.fetch_leads(runner_id, year)` pulls it per attempt straight from Savant's
basestealing-running-game service. This script loops that over every *qualified* runner
league-wide (2023-2025, the Statcast lead-tracking era), aggregates per runner-season,
then ranks the league two ways:

    (a) raw      mean_gain_to_release_ft  — who covers the most ground, full stop.
    (b) residual mean_gain_to_release_ft after regressing out sprint_speed (+ season) —
        ground covered *beyond what the runner's speed predicts* = the timing/jump skill
        that the slow archetype (Naylor / Soto) lives on.

Outputs
-------
  data/DF_GroundCovered_Leaderboard.csv   one row per qualified runner-season
  figures/Fig_GroundCovered_Scatter.png   gain vs sprint, OLS fit, Naylor/Soto annotated
  figures/Fig_GroundCovered_TopN.png      top-25 runner-seasons by speed-adjusted residual
  cv_pilot/discovery/leads_cache/<id>_<Y>.csv   raw per-attempt cache (reruns don't re-hit)

Network: Savant has no sandbox DNS -> run with dangerouslyDisableSandbox.

Usage
-----
  python3 cv_pilot/ground_covered_leaderboard.py
  python3 cv_pilot/ground_covered_leaderboard.py --min-attempts 5 --min-tracked 3
"""
from __future__ import annotations
import argparse
import csv
import sys
import time
from pathlib import Path
from statistics import mean, median

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent          # Statcast Analysis Core/
ROOT = HERE.parent.parent                        # repo root (skip back through Computer Vision/)
DISCOVERY = HERE.parent / "discovery"            # Computer Vision/discovery
CACHE = DISCOVERY / "leads_cache"
DATA = ROOT / "Data Frame"
FIGS = ROOT / "Figures"

# import the native-Statcast leads fetcher (same dir)
sys.path.insert(0, str(HERE))
from fetch_leads import fetch_leads, LEADS_COLS  # noqa: E402

SEASONS = [2023, 2024, 2025, 2026]                 # Statcast lead-tracking era (2026 partial)
NAYLOR_ID, SOTO_ID = 647304, 665742               # highlighted archetype runners
# 2026 is a partial season (~1/3 complete as of May 2026); use lower thresholds
VOL_THRESHOLDS = {2023: 10, 2024: 10, 2025: 10, 2026: 3}   # n_tracked required for ranking


# ----------------------------------------------------------------------------- universe
def load_universe(year: int) -> dict[int, dict]:
    """runner_id -> {name, name_tag, sprint_speed_ftps, sprint_pctile, sb, cs, attempts}."""
    path = DISCOVERY / f"runners_{year}_{year}.csv"
    if not path.exists():
        sys.exit(f"[fatal] missing universe {path} — run discover_runners.py --start {year} --end {year}")
    out: dict[int, dict] = {}
    with open(path, newline="") as f:
        for r in csv.DictReader(f):
            try:
                rid = int(r["runner_id"])
            except (KeyError, TypeError, ValueError):
                continue
            out[rid] = r
    return out


def _f(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


# ----------------------------------------------------------------------------- leads (cached)
def get_leads_cached(runner_id: int, year: int, sleep: float) -> list[dict]:
    """fetch_leads with an on-disk cache so reruns never re-hit Savant."""
    CACHE.mkdir(parents=True, exist_ok=True)
    cf = CACHE / f"{runner_id}_{year}.csv"
    if cf.exists():
        with open(cf, newline="") as f:
            rows = list(csv.DictReader(f))
        for r in rows:                                   # restore numeric types
            for k in ("lead_at_firstmove_ft", "gain_to_release_ft",
                      "lead_at_release_ft", "run_value"):
                r[k] = _f(r.get(k))
        return rows
    rows = fetch_leads(runner_id, year)
    with open(cf, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=LEADS_COLS)
        w.writeheader()
        w.writerows(rows)
    if sleep:
        time.sleep(sleep)                                # polite to Savant
    return rows


def aggregate(runner_id: int, year: int, u: dict, sleep: float) -> dict | None:
    rows = get_leads_cached(runner_id, year, sleep)
    gains = [r["gain_to_release_ft"] for r in rows if r.get("gain_to_release_ft") is not None]
    fm = [r["lead_at_firstmove_ft"] for r in rows if r.get("lead_at_firstmove_ft") is not None]
    rel = [r["lead_at_release_ft"] for r in rows if r.get("lead_at_release_ft") is not None]
    rv = [r["run_value"] for r in rows if r.get("run_value") is not None]
    sb = sum(1 for r in rows if r.get("result") == "SB")
    cs = sum(1 for r in rows if r.get("result") == "CS")
    if not gains:
        return None
    return {
        "runner_id": runner_id,
        "player_name": u.get("name", ""),
        "name_tag": u.get("name_tag", ""),
        "season": year,
        "team": u.get("team", ""),
        "position": u.get("position", ""),
        "sprint_speed_ftps": _f(u.get("sprint_speed_ftps")),
        "sprint_pctile": _f(u.get("sprint_pctile")),
        "n_tracked": len(gains),
        "SB": sb,
        "CS": cs,
        "SB_pct": round(100.0 * sb / (sb + cs), 1) if (sb + cs) else None,
        "mean_gain_to_release_ft": round(mean(gains), 2),
        "median_gain_to_release_ft": round(median(gains), 2),
        "mean_lead_at_firstmove_ft": round(mean(fm), 2) if fm else None,
        "mean_lead_at_release_ft": round(mean(rel), 2) if rel else None,
        "mean_run_value": round(mean(rv), 3) if rv else None,
    }


# ----------------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="League-wide ground-covered (secondary distance) leaderboard")
    ap.add_argument("--min-attempts", type=int, default=5,
                    help="qualify: season SB+CS >= this in the universe (default 5)")
    ap.add_argument("--min-tracked", type=int, default=3,
                    help="require >= this many Statcast-tracked attempts to appear at all (default 3)")
    ap.add_argument("--min-vol", type=int, default=10,
                    help="volume-qualified threshold for the ranked leaderboard (default 10)")
    ap.add_argument("--sleep", type=float, default=0.3, help="polite delay between Savant pulls")
    args = ap.parse_args()

    # ---- pull + aggregate every qualified runner-season -----------------------------
    recs: list[dict] = []
    for year in SEASONS:
        # 2026 is a partial season — use lower att threshold
        yr_min_att = 1 if year == 2026 else args.min_attempts
        uni = load_universe(year)
        qualified = [(rid, u) for rid, u in uni.items()
                     if _f(u.get("attempts")) is not None and _f(u["attempts"]) >= yr_min_att]
        print(f"[{year}] universe={len(uni)}  qualified(SB+CS>={yr_min_att})={len(qualified)}"
              + ("  [partial season — min_att=1]" if year == 2026 else ""))
        for i, (rid, u) in enumerate(qualified, 1):
            try:
                rec = aggregate(rid, year, u, args.sleep)
            except Exception as e:
                print(f"   ! {rid} {year} failed: {e}")
                continue
            if rec is None:
                continue
            recs.append(rec)
            if i % 25 == 0:
                print(f"   .. {i}/{len(qualified)}")

    df = pd.DataFrame(recs)
    print(f"\n[agg] {len(df)} runner-seasons with tracked attempts")

    # keep present sprint speed (speed-adjustment needs it); flag volume-qualified.
    # Thin samples stay in the file but are NOT ranked.
    # 2026 uses a lower n_tracked threshold (partial season; ~1/3 complete as of May 2026).
    df = df[df["sprint_speed_ftps"].notna()].copy()
    # per-season volume threshold
    df["vol_threshold"] = df["season"].map(VOL_THRESHOLDS).fillna(args.min_vol).astype(int)
    df["volume_qualified"] = df["n_tracked"] >= df["vol_threshold"]
    # filter out sub-minimum-tracked rows entirely
    df = df[df["n_tracked"] >= df["season"].map({2026: 1}).fillna(args.min_tracked).astype(int)].copy()
    df["partial_season"] = df["season"] == 2026   # flag for downstream use
    nvq = int(df["volume_qualified"].sum())
    n2026 = int((df["season"] == 2026).sum())
    print(f"[agg] {len(df)} kept (sprint present); "
          f"{nvq} volume-qualified -> ranked  [{n2026} rows from 2026 partial season, threshold=3]")

    # ---- speed-adjusted residual ----------------------------------------------------
    # Fit gain ~ sprint + season on the VOLUME-QUALIFIED subset (reliable per-season
    # means), then score every row against that model. Residual = ground covered beyond
    # what the runner's speed (and era) predicts = the timing/jump skill.
    seasons_sorted = sorted(df["season"].unique())
    Xcols = ["sprint_speed_ftps"]
    for s in seasons_sorted[1:]:                          # drop-first dummy coding
        df[f"is_{s}"] = (df["season"] == s).astype(float)
        Xcols.append(f"is_{s}")
    vq = df[df["volume_qualified"]]
    reg = LinearRegression().fit(vq[Xcols].to_numpy(float),
                                 vq["mean_gain_to_release_ft"].to_numpy(float))
    df["expected_gain_ft"] = np.round(reg.predict(df[Xcols].to_numpy(float)), 2)
    df["gain_residual_ft"] = np.round(df["mean_gain_to_release_ft"] - df["expected_gain_ft"], 2)
    slope_sprint = reg.coef_[0]

    # bivariate (visual) line + correlation on volume-qualified (+ all-row corroboration)
    sp = vq["sprint_speed_ftps"].to_numpy(float)
    gn = vq["mean_gain_to_release_ft"].to_numpy(float)
    b1, b0 = np.polyfit(sp, gn, 1)
    r_sp_gain = float(np.corrcoef(sp, gn)[0, 1])
    r_sp_gain_all = float(np.corrcoef(df["sprint_speed_ftps"], df["mean_gain_to_release_ft"])[0, 1])

    # does ground covered explain steals? runner-season corr with SB% (volume-qualified)
    vqsb = df[df["volume_qualified"] & df["SB_pct"].notna()]
    r_gain_sb = float(np.corrcoef(vqsb["mean_gain_to_release_ft"], vqsb["SB_pct"])[0, 1])
    r_resid_sb = float(np.corrcoef(vqsb["gain_residual_ft"], vqsb["SB_pct"])[0, 1])

    # ---- rankings (only among volume-qualified; thin samples left <NA>) --------------
    df["rank_raw"] = np.nan
    df["rank_residual"] = np.nan
    m = df["volume_qualified"]
    df.loc[m, "rank_raw"] = df.loc[m, "mean_gain_to_release_ft"].rank(ascending=False, method="min")
    df.loc[m, "rank_residual"] = df.loc[m, "gain_residual_ft"].rank(ascending=False, method="min")
    df["pctile_residual"] = (100.0 * (1 - (df["rank_residual"] - 1) / (nvq - 1))).round(1)
    for c in ("rank_raw", "rank_residual"):
        df[c] = df[c].astype("Int64")
    df = df.sort_values(["volume_qualified", "gain_residual_ft"],
                        ascending=[False, False]).reset_index(drop=True)

    # ---- write CSV ------------------------------------------------------------------
    DATA.mkdir(parents=True, exist_ok=True)
    cols = ["rank_residual", "rank_raw", "player_name", "season", "team", "position",
            "sprint_speed_ftps", "sprint_pctile", "volume_qualified", "partial_season",
            "vol_threshold", "n_tracked",
            "SB", "CS", "SB_pct",
            "mean_gain_to_release_ft", "median_gain_to_release_ft",
            "mean_lead_at_firstmove_ft", "mean_lead_at_release_ft", "mean_run_value",
            "expected_gain_ft", "gain_residual_ft", "pctile_residual", "runner_id"]
    # Write "Ground Covered" sheet into consolidated xlsx (preserves other sheets)
    out_xlsx = DATA / "Naylor Blueprint.xlsx"
    if out_xlsx.exists():
        wb = load_workbook(out_xlsx)
        if "Ground Covered" in wb.sheetnames:
            del wb["Ground Covered"]
        wb.save(out_xlsx)
        mode = "a"
    else:
        mode = "w"
    with pd.ExcelWriter(out_xlsx, engine="openpyxl", mode=mode) as writer:
        df[cols].to_excel(writer, sheet_name="Ground Covered", index=False)
    out_csv = out_xlsx  # update reference for print below

    # ---- report --------------------------------------------------------------------
    print("\n" + "=" * 78)
    print(" GROUND-COVERED (secondary distance, first move -> release) LEADERBOARD")
    print("=" * 78)
    print(f"   metric: gain_to_release_ft = r_secondary_lead - r_primary_lead (Statcast)")
    print(f"   volume-qualified ranked: {nvq} (n_tracked>={args.min_vol}); "
          f"total tracked: {len(df)}   seasons: {[int(s) for s in seasons_sorted]}")
    print(f"   OLS gain ~ sprint(+season) on vq: sprint slope = {slope_sprint:+.3f} ft per ft/s")
    print(f"   corr(sprint, gain): vq={r_sp_gain:+.3f}  all={r_sp_gain_all:+.3f}  "
          f"(negative ⇒ slower runners cover MORE ground)")
    print(f"   corr(mean_gain, SB%)     = {r_gain_sb:+.3f}")
    print(f"   corr(gain_residual, SB%) = {r_resid_sb:+.3f}")

    show = ["rank_residual", "player_name", "season", "sprint_speed_ftps", "sprint_pctile",
            "n_tracked", "SB", "CS", "SB_pct", "mean_gain_to_release_ft", "gain_residual_ft"]
    print("\n   Top 20 volume-qualified by speed-adjusted residual (ground beyond speed-expected):")
    with pd.option_context("display.max_columns", None, "display.width", 160):
        print(df[show].head(20).to_string(index=False))

    ns = df[df["runner_id"].isin([NAYLOR_ID, SOTO_ID])]
    print("\n   Naylor & Soto:")
    with pd.option_context("display.max_columns", None, "display.width", 160):
        print(ns[show].to_string(index=False))

    # ---- figures -------------------------------------------------------------------
    FIGS.mkdir(parents=True, exist_ok=True)
    _fig_scatter(df, b0, b1, r_sp_gain, slope_sprint)
    _fig_topn(df, n=25)

    print(f"\n[write] {out_csv}")
    print(f"[write] {FIGS / 'Fig_GroundCovered_Scatter.png'}")
    print(f"[write] {FIGS / 'Fig_GroundCovered_TopN.png'}")

    # one-line machine-readable stats footer (handy for the findings writeup)
    print("\n[stats] " + " ".join([
        f"n_vq={nvq}", f"n_total={len(df)}", f"sprint_slope={slope_sprint:.3f}",
        f"corr_sprint_gain_vq={r_sp_gain:.3f}", f"corr_sprint_gain_all={r_sp_gain_all:.3f}",
        f"corr_gain_sb={r_gain_sb:.3f}", f"corr_resid_sb={r_resid_sb:.3f}",
    ]))


def _annot(df):
    """rows to label on the scatter: Naylor, Soto, plus slow volume-qualified runners."""
    keep = set(df.index[df["runner_id"].isin([NAYLOR_ID, SOTO_ID])])
    # a few of the slowest *volume-qualified* runners with strong residuals, for context
    slow = df[(df["sprint_pctile"] <= 25) & df["volume_qualified"]].sort_values(
        "gain_residual_ft", ascending=False)
    keep.update(slow.head(6).index)
    return keep


def _fig_scatter(df, b0, b1, r_sp_gain, slope_sprint):
    fig, ax = plt.subplots(figsize=(10, 7))
    thin = df[~df["volume_qualified"]]
    vqd = df[df["volume_qualified"]]
    ax.scatter(thin["sprint_speed_ftps"], thin["mean_gain_to_release_ft"],
               s=12, c="#dce6f0", alpha=0.6, edgecolor="none", label="thin sample (<10 tracked)")
    ax.scatter(vqd["sprint_speed_ftps"], vqd["mean_gain_to_release_ft"],
               s=26, c="#9fb3c8", alpha=0.8, edgecolor="none", label="≥10 tracked attempts")
    sp = vqd["sprint_speed_ftps"]
    xs = np.linspace(sp.min(), sp.max(), 100)
    ax.plot(xs, b1 * xs + b0, color="#334e68", lw=2,
            label=f"OLS fit (corr={r_sp_gain:+.2f}; slope={slope_sprint:+.2f} ft per ft/s)")
    for idx in _annot(df):
        r = df.loc[idx]
        star = r["runner_id"] in (NAYLOR_ID, SOTO_ID)
        ax.scatter([r["sprint_speed_ftps"]], [r["mean_gain_to_release_ft"]],
                   s=140 if star else 60, c="#d7263d" if star else "#f4a259",
                   edgecolor="black", zorder=5, marker="*" if star else "o")
        ax.annotate(f"{r['player_name'].split()[-1]} '{str(r['season'])[2:]}",
                    (r["sprint_speed_ftps"], r["mean_gain_to_release_ft"]),
                    textcoords="offset points", xytext=(7, 4),
                    fontsize=9, fontweight="bold" if star else "normal",
                    color="#d7263d" if star else "#7a4f01")
    ax.set_xlabel("Sprint speed (ft/s)  —  lower = slower runner", fontsize=11)
    ax.set_ylabel("Mean ground covered, first move → release (ft)", fontsize=11)
    ax.set_title("Stealing on secondary distance, not speed\n"
                 "Naylor & Soto cover elite ground despite bottom-percentile sprint speed",
                 fontsize=12, fontweight="bold")
    ax.legend(loc="upper left", fontsize=9)
    ax.grid(alpha=0.25)
    fig.tight_layout()
    fig.savefig(FIGS / "Fig_GroundCovered_Scatter.png", dpi=130)
    plt.close(fig)


def _fig_topn(df, n=25):
    top = (df[df["volume_qualified"]]
           .sort_values("gain_residual_ft", ascending=False).head(n).iloc[::-1])
    labels = [f"{r.player_name.split()[-1]} '{str(r.season)[2:]}" for r in top.itertuples()]
    colors = ["#d7263d" if rid in (NAYLOR_ID, SOTO_ID) else "#3e7cb1"
              for rid in top["runner_id"]]
    fig, ax = plt.subplots(figsize=(9, 9))
    ax.barh(range(len(top)), top["gain_residual_ft"], color=colors)
    ax.set_yticks(range(len(top)))
    ax.set_yticklabels(labels, fontsize=9)
    for i, (v, sp) in enumerate(zip(top["gain_residual_ft"], top["sprint_speed_ftps"])):
        ax.text(v + 0.02, i, f"{v:+.2f}  ({sp:.1f} ft/s)", va="center", fontsize=8)
    ax.axvline(0, color="black", lw=0.8)
    ax.set_xlabel("Ground covered ABOVE speed-expected (ft)  —  the timing/jump skill", fontsize=11)
    ax.set_title(f"Top {n} runner-seasons: secondary distance beyond what sprint speed predicts\n"
                 "(red = Naylor / Soto)", fontsize=12, fontweight="bold")
    ax.grid(axis="x", alpha=0.25)
    fig.tight_layout()
    fig.savefig(FIGS / "Fig_GroundCovered_TopN.png", dpi=130)
    plt.close(fig)


if __name__ == "__main__":
    main()
